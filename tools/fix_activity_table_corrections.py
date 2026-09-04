# -*- coding: utf-8 -*-
"""按 2026-09-04 校对结果修正 王晰演出活动.xlsx：
1. 合并/去重 2025-12-06 CMG童声歌会
2. 合并/去重 2026-02-12 光从东方来/长安夜色
3. 删除 2026-08-24 广州歌声相逢（错误）
4. 补充 2025-12-31 启航2026 地点/演唱歌曲
5. 2017-11 专辑信息修正见 data/timeline.json（已直接修改）

用法：
  python -X utf8 D:\\wx409.github.io\\tools\\fix_activity_table_corrections.py --dry-run
  python -X utf8 D:\\wx409.github.io\\tools\\fix_activity_table_corrections.py
"""
from __future__ import annotations

import argparse
import datetime
import shutil
from pathlib import Path

import openpyxl

XLSX = Path(r"E:\wx\index_records\王晰演出活动.xlsx")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()

    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active

    rows = list(ws.iter_rows())
    # build row index by name
    def find(name_part: str):
        out = []
        for row in rows:
            name = str(row[1].value or "")
            if name_part in name:
                out.append(row[0].row)
        return out

    cmg_main_rows = find("总台《CMG童声歌会》播出")
    cmg_dup_rows = find("王晰助阵CMG童声歌会")
    gz_dup_rows = find("《光从东方来》视听盛会")
    gz_sub_rows = find("王晰长安夜色低音演出")
    err_rows = find("广州歌声相逢")
    qh_rows = find("《启航2026》总台跨年晚会")

    print("CMG main", cmg_main_rows, "dup", cmg_dup_rows)
    print("GuangCong main", gz_dup_rows, "sub", gz_sub_rows)
    print("error", err_rows)
    print("Qihang", qh_rows)

    # 1) 合并 CMG：主行备注补充频道，删除重复行
    if cmg_main_rows:
        r = cmg_main_rows[0]
        note = ws.cell(row=r, column=5).value or ""
        if "CCTV-14少儿频道" not in note:
            ws.cell(row=r, column=5).value = (note + "；CCTV-14少儿频道").strip("；")
            print(f"[merge] row {r} 备注补 CCTV-14少儿频道")

    # 2) 合并光从东方来：主行备注补充长安夜色说明，删除副行
    if gz_dup_rows:
        r = gz_dup_rows[0]
        note = ws.cell(row=r, column=5).value or ""
        if "西安夜色" not in note:
            ws.cell(row=r, column=5).value = (note + "；王晰在西安夜色中用低音演绎启新生主题").strip("；")
            print(f"[merge] row {r} 备注补 长安夜色说明")

    # 3) 启航2026：补地点/歌曲
    if qh_rows:
        r = qh_rows[0]
        loc = ws.cell(row=r, column=4).value
        song = ws.cell(row=r, column=6).value
        if loc is None or str(loc).strip() in ("", "—"):
            ws.cell(row=r, column=4).value = "山西吕梁新区体育中心"
            print(f"[update] row {r} 地点 -> 山西吕梁新区体育中心")
        if song is None or not str(song).strip():
            ws.cell(row=r, column=6).value = "在路上"
            print(f"[update] row {r} 演唱曲目 -> 在路上")

    # 4) 删除重复/错误行
    delete_rows = []
    if cmg_dup_rows:
        delete_rows.append(cmg_dup_rows[0])
    if gz_sub_rows:
        delete_rows.append(gz_sub_rows[0])
    if err_rows:
        delete_rows.append(err_rows[0])
    delete_rows = sorted(set(delete_rows), reverse=True)
    if delete_rows:
        print("[delete rows]", delete_rows)
        if not a.dry_run:
            for r in delete_rows:
                ws.delete_rows(r)
            # 序号列可能因删除需要重新整理；不自动改，保留后续人工/生成器重排

    if a.dry_run:
        print("\n[dry-run] 以上为预览，未修改。")
        return 0

    bak = XLSX.with_name(
        XLSX.stem + f"_bak_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    shutil.copy2(XLSX, bak)
    wb.save(XLSX)
    print(f"[备份] {bak.name}")
    print(f"[已保存] {XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
